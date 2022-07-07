from openpyxl import load_workbook
from functools import reduce

file_location = 'main.xlsx'
wb = load_workbook(file_location)

set_list = []
for i in range(1, 3):
    wb.active = i
    set_list.append(set(map(lambda t: t.value, wb.active['A'][1:])))

set_list = reduce(lambda t, q: t & q, set_list)

wb.active = 0
for cell in wb.active['A'][1:]:
    cell.value = None

for ind, val in enumerate(set_list):
    wb.active.cell(row=ind+2, column=1, value=val)

wb.save(file_location)
